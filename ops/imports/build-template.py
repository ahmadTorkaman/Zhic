"""
Build the mass-data import template for Zhic.

Sheets:
  1. Products   — strictly-required fields only (Option A)
  2. Showrooms  — strictly-required fields only (Option A)
  3. Reference  — valid slugs and enum values

Source of truth: services/api/src/collections/{Products,Showrooms}.ts

Run:
  python3 ops/imports/build-template.py
Output:
  ops/imports/zhic-mass-data.xlsx
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).parent / "zhic-mass-data.xlsx"

# --- shared styling ---------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="5F7760")  # Zhic forest
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
EXAMPLE_FILL = PatternFill("solid", fgColor="F4EFE7")  # warm paper
EXAMPLE_FONT = Font(italic=True, color="6B6258")
REF_HEADER_FILL = PatternFill("solid", fgColor="C49A6C")  # caramel
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def style_example_row(ws, ncols: int, row: int = 2) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = EXAMPLE_FILL
        cell.font = EXAMPLE_FONT


# --- reference data (must match seed.ts) ------------------------------------

DESIGN_SLUGS = ["aramesh", "bahar"]
MATERIAL_SLUGS = ["walnut", "beech", "oak", "belgian-linen", "velvet"]
AVAILABILITY_VALUES = ["in_stock", "made_to_order", "backorder", "discontinued"]
STATUS_VALUES = ["draft", "published"]

DESIGN_LABELS = {"aramesh": "آرامش", "bahar": "بهار"}
MATERIAL_LABELS = {
    "walnut": "چوب گردو",
    "beech": "راش",
    "oak": "بلوط",
    "belgian-linen": "کتان بلژیکی",
    "velvet": "مخمل",
}
AVAILABILITY_LABELS = {
    "in_stock": "موجود",
    "made_to_order": "ساخت به‌سفارش",
    "backorder": "در انتظار",
    "discontinued": "ناموجود",
}
STATUS_LABELS = {"draft": "پیش‌نویس", "published": "منتشرشده"}


# --- workbook ---------------------------------------------------------------

wb = Workbook()

# =============================================================================
# Sheet 1 — Products
# =============================================================================
ws_p = wb.active
ws_p.title = "Products"

product_columns = [
    (
        "name",
        "نام محصول",
        "تخت آرامش گردو",
        "Required. Persian product name.",
    ),
    (
        "design",
        "طرح (slug)",
        "aramesh",
        "Required. One slug from Reference sheet: aramesh, bahar.",
    ),
    (
        "materialIds",
        "متریال‌ها (slugs, comma-separated)",
        "walnut, belgian-linen",
        "Required. One or more material slugs separated by commas. "
        "Valid: walnut, beech, oak, belgian-linen, velvet.",
    ),
    (
        "sku",
        "SKU",
        "BED-001",
        "Required. Pattern AAA-NNN (2–4 capital letters, dash, 3–5 digits). Must be unique.",
    ),
    (
        "basePriceTomans",
        "قیمت پایه (تومان)",
        85000000,
        "Required. Integer tomans (NOT rials). Will be ×10 at import to store as rials.",
    ),
    (
        "availability",
        "وضعیت موجودی",
        "made_to_order",
        "Required. One of: in_stock, made_to_order, backorder, discontinued.",
    ),
    (
        "leadTimeDays",
        "زمان تحویل (روز)",
        56,
        "Required. Integer days. Default for made-to-order is 56.",
    ),
    (
        "status",
        "وضعیت",
        "draft",
        "Required. draft (پیش‌نویس) or published (منتشرشده). Use draft until reviewed.",
    ),
]

# header
for i, (key, label, _example, doc) in enumerate(product_columns, start=1):
    cell = ws_p.cell(row=1, column=i, value=key)
    cell.comment = Comment(f"{label}\n\n{doc}", "Zhic")

# Persian sub-header (row 2 hidden? no — keep visible as a label hint).
# We use row 2 for the example row instead so it stays visible.
for i, (_key, _label, example, _doc) in enumerate(product_columns, start=1):
    ws_p.cell(row=2, column=i, value=example)

style_header(ws_p, len(product_columns))
style_example_row(ws_p, len(product_columns), row=2)

# column widths
widths_p = [26, 18, 38, 12, 18, 18, 14, 14]
for i, w in enumerate(widths_p, start=1):
    ws_p.column_dimensions[get_column_letter(i)].width = w

# data validations (rows 2 → 1000)
dv_design = DataValidation(
    type="list",
    formula1=f'"{",".join(DESIGN_SLUGS)}"',
    allow_blank=False,
    showDropDown=False,  # showDropDown=False means dropdown IS shown (openpyxl quirk)
)
dv_design.error = "Design must be one of: " + ", ".join(DESIGN_SLUGS)
dv_design.errorTitle = "Invalid design slug"
dv_design.add("B2:B1000")
ws_p.add_data_validation(dv_design)

dv_avail = DataValidation(
    type="list",
    formula1=f'"{",".join(AVAILABILITY_VALUES)}"',
    allow_blank=False,
    showDropDown=False,
)
dv_avail.error = "Availability must be one of: " + ", ".join(AVAILABILITY_VALUES)
dv_avail.errorTitle = "Invalid availability"
dv_avail.add("F2:F1000")
ws_p.add_data_validation(dv_avail)

dv_status = DataValidation(
    type="list",
    formula1=f'"{",".join(STATUS_VALUES)}"',
    allow_blank=False,
    showDropDown=False,
)
dv_status.error = "Status must be: draft or published"
dv_status.errorTitle = "Invalid status"
dv_status.add("H2:H1000")
ws_p.add_data_validation(dv_status)

dv_lead = DataValidation(
    type="whole",
    operator="greaterThanOrEqual",
    formula1=0,
    allow_blank=False,
)
dv_lead.error = "Lead time must be a non-negative integer (days)"
dv_lead.errorTitle = "Invalid lead time"
dv_lead.add("G2:G1000")
ws_p.add_data_validation(dv_lead)

dv_price = DataValidation(
    type="whole",
    operator="greaterThanOrEqual",
    formula1=0,
    allow_blank=False,
)
dv_price.error = "Price must be a non-negative integer in tomans"
dv_price.errorTitle = "Invalid price"
dv_price.add("E2:E1000")
ws_p.add_data_validation(dv_price)

# format price column with thousands separator
for row in range(2, 1001):
    ws_p.cell(row=row, column=5).number_format = "#,##0"

# =============================================================================
# Sheet 2 — Showrooms
# =============================================================================
ws_s = wb.create_sheet("Showrooms")

showroom_columns = [
    (
        "name",
        "نام شوروم",
        "شوروم تهران",
        "Required. Showroom display name (Persian).",
    ),
    (
        "city",
        "شهر",
        "تهران",
        "Required. City name in Persian (e.g., تهران, اصفهان, همدان).",
    ),
    (
        "phone",
        "تلفن (عمومی)",
        "021-88123456",
        "Required. Public phone number. Format: 021-88123456 or +982188123456.",
    ),
]

for i, (key, label, _example, doc) in enumerate(showroom_columns, start=1):
    cell = ws_s.cell(row=1, column=i, value=key)
    cell.comment = Comment(f"{label}\n\n{doc}", "Zhic")

for i, (_key, _label, example, _doc) in enumerate(showroom_columns, start=1):
    ws_s.cell(row=2, column=i, value=example)

style_header(ws_s, len(showroom_columns))
style_example_row(ws_s, len(showroom_columns), row=2)

widths_s = [26, 18, 22]
for i, w in enumerate(widths_s, start=1):
    ws_s.column_dimensions[get_column_letter(i)].width = w

# =============================================================================
# Sheet 3 — Reference
# =============================================================================
ws_r = wb.create_sheet("Reference")

ws_r["A1"] = "Reference values — copy slugs into Products / Showrooms sheets"
ws_r["A1"].font = Font(bold=True, italic=True, color="3D2F26", size=12)
ws_r.merge_cells("A1:C1")
ws_r.row_dimensions[1].height = 24


def write_ref_block(ws, start_row: int, title: str, items: list[tuple[str, str]]) -> int:
    """Returns next-available row."""
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=start_row, column=1).fill = REF_HEADER_FILL
    ws.cell(row=start_row, column=2, value="Persian label").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=start_row, column=2).fill = REF_HEADER_FILL
    for r, (slug, label) in enumerate(items, start=start_row + 1):
        ws.cell(row=r, column=1, value=slug).font = Font(name="Consolas")
        ws.cell(row=r, column=2, value=label)
    return start_row + 1 + len(items) + 1  # blank line spacer


next_row = 3
next_row = write_ref_block(
    ws_r, next_row, "Designs (column: design)",
    [(s, DESIGN_LABELS[s]) for s in DESIGN_SLUGS],
)
next_row = write_ref_block(
    ws_r, next_row, "Materials (column: materialIds — comma-separate multiples)",
    [(s, MATERIAL_LABELS[s]) for s in MATERIAL_SLUGS],
)
next_row = write_ref_block(
    ws_r, next_row, "Availability (column: availability)",
    [(s, AVAILABILITY_LABELS[s]) for s in AVAILABILITY_VALUES],
)
next_row = write_ref_block(
    ws_r, next_row, "Status (column: status)",
    [(s, STATUS_LABELS[s]) for s in STATUS_VALUES],
)

# notes block
notes = [
    "",
    "Notes:",
    "• SKU pattern: AAA-NNN (e.g., BED-001, NS-042). Must be unique across all products.",
    "• Price: enter in TOMANS as integer. Will be multiplied by 10 to store as rials.",
    "• materialIds: at least one slug required; separate multiples with commas (e.g., walnut, oak).",
    "• Showroom city: Persian text. Used for SMS routing — match exactly across showrooms.",
    "• Row 2 in each sheet is an EXAMPLE row. Replace it or delete before sending the file back.",
    "• Add as many rows as you need — validation extends to row 1000.",
]
for line in notes:
    ws_r.cell(row=next_row, column=1, value=line)
    if line.startswith("Notes"):
        ws_r.cell(row=next_row, column=1).font = Font(bold=True)
    next_row += 1

ws_r.column_dimensions["A"].width = 30
ws_r.column_dimensions["B"].width = 24
ws_r.column_dimensions["C"].width = 50

# --- save -------------------------------------------------------------------
wb.save(OUT)
print(f"Wrote {OUT}")
